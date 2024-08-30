import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from ttkthemes import ThemedTk  # Для тем оформления
from tkinter import ttk
import threading  # Для работы с потоками


class FileProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("OzonTo1C")
        self.root.geometry("600x500")
        self.root.iconbitmap("free-icon-convert-4700974.ico")  # Замените на свой значок

        # Фон и тема
        self.root.set_theme("arc")  # Применяем стиль "arc"

        # Текстовые переменные для путей
        self.input_path = tk.StringVar()
        self.ids_file_path = tk.StringVar(value='Z:\\ozon программа\\Копия УИД.xlsx')  # Путь к файлу с ID

        # Размещение кнопок и логов
        self.create_widgets()

        self.ozon_data = None

    def create_widgets(self):
        # Кнопка загрузки файла с Ozon
        self.upload_button = ttk.Button(self.root, text="Загрузить файл с ozon", command=self.load_ozon_file)
        self.upload_button.pack(fill="x", padx=20, pady=10)

        # Кнопка сохранения файла для 1С
        self.save_button = ttk.Button(self.root, text="Сохранить файл для 1С", command=self.save_for_1c)
        self.save_button.pack(fill="x", padx=20, pady=10)

        # Поле ввода пути к файлу с UID
        self.ids_path_entry = ttk.Entry(self.root, textvariable=self.ids_file_path)
        self.ids_path_entry.pack(fill="x", padx=20, pady=10)

        # Виджет для логов
        self.log_widget = tk.Text(self.root, height=10, state="disabled")
        self.log_widget.pack(fill="both", padx=20, pady=20, expand=True)

    def log_message(self, message):
        self.log_widget.configure(state="normal")
        self.log_widget.insert(tk.END, message + '\n')
        self.log_widget.configure(state="disabled")
        self.log_widget.see("end")

    def load_ozon_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if file_path:
            try:
                # Загрузка данных с группировкой по артикулу и суммированием количества
                self.ozon_data = pd.read_csv(file_path, delimiter=';', on_bad_lines='skip')
                self.ozon_data = self.ozon_data.groupby('Артикул', as_index=False)['Количество'].sum()
                self.input_path.set(file_path)
                self.log_message(f"Файл {file_path} успешно загружен.")
                messagebox.showinfo("Информация", "Файл успешно загружен!")
            except Exception as e:
                self.log_message(f"Ошибка загрузки файла: {e}")
                messagebox.showerror("Ошибка", f"Не удалось загрузить файл: {e}")

    def save_for_1c(self):
        if self.ozon_data is None:
            messagebox.showwarning("Предупреждение", "Необходимо сначала загрузить файл с ozon.")
            return

        ids_file_path = self.ids_file_path.get()

        # Запуск процесса сохранения в отдельном потоке
        save_thread = threading.Thread(target=self.process_and_save, args=(ids_file_path,))
        save_thread.start()

    def process_and_save(self, ids_file_path):
        try:
            # Чтение файла с ID
            ids_data = pd.read_excel(ids_file_path)

            # Проверка наличия нужных столбцов в обоих файлах
            if 'Артикул' not in self.ozon_data.columns or 'Количество' not in self.ozon_data.columns:
                error_message = "Файл с ozon должен содержать столбцы 'Артикул' и 'Количество'."
                self.log_message(error_message)
                messagebox.showerror("Ошибка", error_message)
                return

            if 'Артикул' not in ids_data.columns or 'UID' not in ids_data.columns:
                error_message = "Файл ids должен содержать столбцы 'Артикул' и 'UID'."
                self.log_message(error_message)
                messagebox.showerror("Ошибка", error_message)
                return

            # Объединение данных по Артикул
            merged_data = pd.merge(self.ozon_data[['Артикул', 'Количество']], ids_data[['Артикул', 'UID']],
                                   on='Артикул', how='left')

            # Проверка на отсутствие UID и логирование
            missing_uids = merged_data[merged_data['UID'].isna()]['Артикул'].tolist()
            if missing_uids:
                for missing_art in missing_uids:
                    self.log_message(f"Артикул {missing_art} не найден в файле с UID.")

            # Удаление строк с отсутствующими UID для сохранения в файл
            result_df = merged_data[['UID', 'Количество']].dropna(subset=['UID'])

            # Сохранение в файл на основе шаблона
            template_path = 'template.xlsx'
            wb = load_workbook(template_path)
            ws = wb.active

            # Запись UID и Количество в нужные столбцы
            for idx, (uid, qty) in enumerate(zip(result_df['UID'], result_df['Количество']), start=11):
                ws.cell(row=idx, column=4, value=uid)
                ws.cell(row=idx, column=18, value=qty)

            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if save_path:
                wb.save(save_path)
                self.log_message(f"Файл успешно сохранен как {save_path}")
                messagebox.showinfo("Информация", "Файл успешно сохранен!")

        except Exception as e:
            error_message = f"Ошибка сохранения файла: {e}"
            self.log_message(error_message)
            messagebox.showerror("Ошибка", error_message)


if __name__ == "__main__":
    root = ThemedTk(theme="arc")  # Используем тематический виджет для улучшенного дизайна
    app = FileProcessorApp(root)
    root.mainloop()
